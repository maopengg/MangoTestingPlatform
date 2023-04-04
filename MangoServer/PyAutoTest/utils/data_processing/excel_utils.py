# -*- coding: utf-8 -*-
# @Project: auto_test
# @Description: excel文件操作类
# @Time   : 2022-11-04 22:05
# @Author : 毛鹏

import openpyxl
from PyAutoTest.utils.file_data_handle.config_file.config_ini import Config


class ExcelHandler:
    def __init__(self, file):
        self.file = file
        self.config = Config()

    def open_excel(self, sheet_name):
        """打开Excel、获取sheet"""
        wb = openpyxl.load_workbook(self.file)
        # 获取sheet_name
        sheet = wb[sheet_name]
        return sheet

    def get_header(self, sheet_name):
        """获取header(表头)"""
        wb = self.open_excel(sheet_name)
        header = []
        # 遍历第一行
        for i in wb[1]:
            # 将遍历出来的表头字段加入列表
            header.append(i.value)
        return header

    def read_excel(self, sheet_name, path):
        """读取所有数据"""
        global data_dict
        sheet = self.open_excel(sheet_name)
        rows = list(sheet.rows)
        data = []
        # 遍历从第二行开始的每一行数据
        for row in rows[1:]:
            row_data = []
            # 遍历每一行的每个单元格
            for cell in row:
                row_data.append(cell.value)
                # 通过zip函数将两个列表合并成字典
                data_dict = dict(zip(self.get_header(sheet_name), row_data))
            # url拼接
            url = data_dict["url"]
            client = data_dict["client"]
            host = self.config.read_ini(path, client)
            data_dict["url"] = host + url
            data.append(data_dict)
        return data

    @staticmethod
    def write_excel(file, sheet_name, row, cloumn, data):
        """Excel写入数据"""
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        sheet.cell(row, cloumn).value = data
        wb.save(file)
        wb.close()


if __name__ == "__main__":
    # 以下为测试代码
    r = ExcelHandler(r"/case_run\api-tset_case.xlsx")
    ptah = r'D:\Test_Code\Zshop_Api_Auto\pytest.ini'
    w = r.read_excel('tset_case', ptah)
    for i in w:
        print(i)
        # # 请在引号内配置用例文件路径
        # self.case_list = ExcelHandler(r"D:\Test_Code\Zshop_Api_Auto\case_run\api-tset_case.xlsx")
        # # 请在引号内配置执行用例的表
        # self.sheet = "tset_case"
        # # 请在引号内配置配置文件路径
        # self.config = r'D:\Test_Code\Zshop_Api_Auto\config\ui_config.ini'
