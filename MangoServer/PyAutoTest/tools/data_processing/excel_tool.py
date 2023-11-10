# -*- coding: utf-8 -*-
# @Project: auto_test
# @Description:
# @Time   : 2023-03-07 8:24
# @Author : 毛鹏

import openpyxl


class ExcelTools:
    """Excel处理类"""

    @classmethod
    def excel_write(cls, content: bytes, file_path: str):
        """
        查询账户是否存在
        :return: 响应结果，请求url，请求头
        """
        # 以二进制模式打开文件
        with open(file_path, 'wb') as file:
            # 将数据流写入文件
            file.write(content)

    @classmethod
    def excel_rows(cls, file_path: str) -> int:
        """
        查询账户是否存在
        :return: 响应结果，请求url，请求头
        """
        # 以二进制模式打开文件
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        return sheet.max_row

    @classmethod
    def excel_columns(cls, file_path: str) -> int:
        """
        查询账户是否存在
        :return: 响应结果，请求url，请求头
        """
        # 以二进制模式打开文件
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        return sheet.max_column
