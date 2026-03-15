# -*- coding: utf-8 -*-
# @Project: 芒果测试平台
# @Description: AI生成测试用例视图
# @Author : 毛鹏
import io

from django.http import HttpResponse
from rest_framework import serializers
from rest_framework.decorators import action
from rest_framework.request import Request
from rest_framework.viewsets import ViewSet

from src.auto_test.auto_ai.models import AiTestCase
from src.auto_test.auto_ai.views.ai_requirement import AiRequirementSerializers
from src.auto_test.auto_ai.views.ai_requirement_split import AiRequirementSplitSerializers
from src.auto_test.auto_ai.views.ai_test_point import AiTestPointSerializers
from src.auto_test.auto_system.views.product_module import ProductModuleSerializers
from src.auto_test.auto_system.views.project_product import ProjectProductSerializersC
from src.enums.ai_enum import (
    AiCasePriorityEnum, AiCaseTypeEnum, AiCaseTestResultEnum, AiCaseAutoTagEnum
)
from src.tools.decorator.error_response import error_response
from src.tools.view.model_crud import ModelCRUD
from src.tools.view.response_data import ResponseData
from src.tools.view.response_msg import *


class AiTestCaseSerializers(serializers.ModelSerializer):
    create_time = serializers.DateTimeField(format='%Y-%m-%d %H:%M:%S', read_only=True)
    update_time = serializers.DateTimeField(format='%Y-%m-%d %H:%M:%S', read_only=True)

    class Meta:
        model = AiTestCase
        fields = '__all__'


class AiTestCaseSerializersC(serializers.ModelSerializer):
    create_time = serializers.DateTimeField(format='%Y-%m-%d %H:%M:%S', read_only=True)
    update_time = serializers.DateTimeField(format='%Y-%m-%d %H:%M:%S', read_only=True)
    project_product = ProjectProductSerializersC(read_only=True)
    module = ProductModuleSerializers(read_only=True)
    requirement = AiRequirementSerializers(read_only=True)
    requirement_split = AiRequirementSplitSerializers(read_only=True)
    test_point = AiTestPointSerializers(read_only=True)

    class Meta:
        model = AiTestCase
        fields = '__all__'

    @staticmethod
    def setup_eager_loading(queryset):
        queryset = queryset.select_related(
            'project_product',
            'module',
            'requirement',
            'requirement_split',
            'test_point',
        )
        return queryset


class AiTestCaseCRUD(ModelCRUD):
    model = AiTestCase
    queryset = AiTestCase.objects.all()
    serializer_class = AiTestCaseSerializersC
    serializer = AiTestCaseSerializers


class AiTestCaseViews(ViewSet):
    model = AiTestCase
    serializer_class = AiTestCaseSerializers

    @action(methods=['get'], detail=False)
    @error_response('system')
    def by_requirement(self, request: Request):
        """按需求ID获取全部用例"""
        requirement_id = request.query_params.get('requirement_id')
        qs = AiTestCase.objects.filter(requirement_id=requirement_id)
        try:
            qs = AiTestCaseSerializersC.setup_eager_loading(qs)
        except Exception:
            pass
        data = AiTestCaseSerializersC(instance=qs, many=True).data
        return ResponseData.success(RESPONSE_MSG_0001, data=data, value=(qs.count(),))

    @action(methods=['get'], detail=False)
    @error_response('system')
    def by_test_point(self, request: Request):
        """按测试点ID获取用例列表"""
        test_point_id = request.query_params.get('test_point_id')
        qs = AiTestCase.objects.filter(test_point_id=test_point_id)
        try:
            qs = AiTestCaseSerializersC.setup_eager_loading(qs)
        except Exception:
            pass
        data = AiTestCaseSerializersC(instance=qs, many=True).data
        return ResponseData.success(RESPONSE_MSG_0001, data=data, value=(qs.count(),))

    @action(methods=['get'], detail=False)
    @error_response('system')
    def export_excel(self, request: Request):
        """导出测试用例为 Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        requirement_id = request.query_params.get('requirement_id')
        qs = AiTestCase.objects.filter(requirement_id=requirement_id).select_related(
            'requirement', 'requirement_split', 'test_point'
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '功能测试用例'

        headers = [
            '用例编号', '所属模块', '用例标题', '用例类型', '优先级',
            '版本编号', '前置条件', '测试步骤', '预期结果',
            '开发自测', '测试结果', '预发结果', '自动化标识', '备注',
        ]
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        col_widths = [12, 20, 40, 12, 10, 12, 30, 50, 40, 12, 12, 12, 14, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        priority_map = AiCasePriorityEnum.obj()
        case_type_map = AiCaseTypeEnum.obj()
        result_map = AiCaseTestResultEnum.obj()
        auto_tag_map = AiCaseAutoTagEnum.obj()

        for row_idx, case in enumerate(qs, 2):
            steps_text = ''
            if case.steps:
                steps_text = '\n'.join(
                    f"{i + 1}. {s}" if isinstance(s, str)
                    else f"{i + 1}. {s.get('step', '')}"
                    for i, s in enumerate(case.steps)
                )
            split_name = case.requirement_split.name if case.requirement_split else ''
            row_data = [
                case.case_no or '',
                case.module_name or split_name,
                case.title,
                case_type_map.get(case.case_type, ''),
                priority_map.get(case.priority, ''),
                case.version or '',
                case.precondition or '',
                steps_text,
                case.expected,
                result_map.get(case.dev_test_result, ''),
                result_map.get(case.test_result, ''),
                result_map.get(case.pre_release_result, ''),
                auto_tag_map.get(case.auto_tag, ''),
                case.remark or '',
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        req_name = 'cases'
        if requirement_id:
            from src.auto_test.auto_ai.models import AiRequirement
            try:
                req_name = AiRequirement.objects.get(id=requirement_id).name
            except AiRequirement.DoesNotExist:
                pass

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f"attachment; filename*=UTF-8''{req_name.encode('utf-8').hex()}.xlsx"
        )
        return response
